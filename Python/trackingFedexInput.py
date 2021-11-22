#! python3
#Opens several Google search results.
import openpyxl
import requests, sys, webbrowser, bs4 as bs
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
import time
import subprocess

header = ['Tracking Number','Service','Reference','MASTER TRACKING NUMBER','Weight','Delivered to','Dimensions','Total Pieces','Terms','PO Number','Department Number','ship date','Invoice Number','Packaging','Standard Transit','Actual Delivery','total shipment weight','SHIPPER REFERENCE','SPECIAL HANDLING SECTION']
trackingNumbers = []

#Initialize Browser Driver
options = Options()
options.headless = True
options.binary_location = 'C:\\Program Files\\Mozilla Firefox\\firefox.exe'
service = Service('C:\\Users\\garrettp\Downloads\\geckodriver-v0.30.0-win64\\geckodriver.exe')

input_string = input('Enter your list of tracking numbers: ')
print("\n")
trackingNumbers = input_string.split()

currentRow = 2

#Open and initialize workbook
wb = openpyxl.Workbook()
sheet = wb.active

#writing header row to spreadsheet
for y in range(0, len(header)):
    yy = y+1
    sheet.cell(row=1 ,column=yy ).value = header[y]
    y+=1

for z in range(0,len(trackingNumbers)):
    driver = webdriver.Firefox(options=options, service = service)
    number = trackingNumbers[z]

    print("https://www.fedex.com/fedextrack/?trknbr=" + number)
    #URL
    driver.get("https://www.fedex.com/fedextrack/?trknbr=" + number)

    
    # wait for the page to load
    time.sleep(3)

    try:
        shipmentFacts = driver.find_element(By.ID, "tab-list-item-tab_panel_3")

    except:
        print(number + ' Is not found')
        driver.close()
        continue
    shipmentFacts.click()

    # get the page source
    page_source = driver.page_source

    driver.close()

    # parse the HTML
    soup = bs.BeautifulSoup(page_source, "html.parser")

    for x in range(0,19):
        
        b = x +1
        fig = "fig"
        #print(soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text())
        try:
            fig = soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > div:nth-child(2)')
            
            if soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Tracking Number':
                sheet.cell(row=currentRow ,column=1 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Service':
                sheet.cell(row=currentRow ,column=2 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Reference':
                sheet.cell(row=currentRow ,column=3 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Master tracking number':
                sheet.cell(row=currentRow ,column=4 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Weight':
                sheet.cell(row=currentRow ,column=5 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Delivered To':
                sheet.cell(row=currentRow ,column=6 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Dimensions':
                sheet.cell(row=currentRow ,column=7 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Total pieces':
                sheet.cell(row=currentRow ,column=8 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Terms':
                sheet.cell(row=currentRow ,column=9 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Purchase order number':
                sheet.cell(row=currentRow ,column=10 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Department number':
                sheet.cell(row=currentRow ,column=11 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Ship date':
                sheet.cell(row=currentRow ,column=12 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Invoice number': #li.key-value-list__item:nth-child(8) > h4:nth-child(1)  li.key-value-list__item:nth-child(8) > div:nth-child(2)
                sheet.cell(row=currentRow ,column=13 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Packaging':
                sheet.cell(row=currentRow ,column=14 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Standard transit':
                sheet.cell(row=currentRow ,column=15 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Actual delivery':
                sheet.cell(row=currentRow ,column=16 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Total shipment weight':
                sheet.cell(row=currentRow ,column=17 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Shipper reference':
                sheet.cell(row=currentRow ,column=18 ).value = fig[0].get_text()
                x+=1
            elif soup.select('li.key-value-list__item:nth-child(' + str(b) + ') > h4:nth-child(1)')[0].get_text() == 'Special handling section':
                sheet.cell(row=currentRow ,column=19 ).value = fig[0].get_text()
                x+=1
        except: 
            break
        
    currentRow+=1
 
wb.save('fedexTracking.xlsx')

subprocess.call(['C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE','fedexTracking.xlsx'])